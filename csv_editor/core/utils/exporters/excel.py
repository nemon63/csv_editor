from openpyxl import Workbook
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import Qt

class ExcelExporter:
    def __init__(self, model):
        self.model = model

    def export(self, filename):
        try:
            wb = Workbook()
            ws = wb.active
            
            # Заголовки
            for col in range(self.model.columnCount()):
                ws.cell(row=1, column=col+1, 
                       value=self.model.headerData(col, Qt.Horizontal))
            
            # Данные
            for row in range(self.model.rowCount()):
                for col in range(self.model.columnCount()):
                    ws.cell(row=row+2, column=col+1, 
                           value=self.model.data(self.model.index(row, col)))
            
            wb.save(filename)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Excel Export Error", str(e))
            return False

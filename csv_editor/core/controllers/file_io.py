import csv
import sqlite3
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from PyQt5.QtCore import QObject, pyqtSignal, Qt
from PyQt5.QtGui import QStandardItem

class FileIOController(QObject):
    file_loaded = pyqtSignal(list)
    file_saved = pyqtSignal(bool)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file = None
        self.model = None

    def open_file(self, parent_widget, file_type=None):
        """Универсальный метод открытия файлов"""
        if file_type == 'csv':
            filters = "CSV Files (*.csv)"
        elif file_type == 'excel':
            filters = "Excel Files (*.xlsx *.xls)"
        else:
            filters = "All Files (*);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)"

        filename, _ = QFileDialog.getOpenFileName(
            parent_widget,
            "Open File",
            "",
            filters
        )

        if not filename:
            return False

        self.current_file = filename

        try:
            if filename.endswith('.csv'):
                data = self._load_csv_file(filename)
            else:
                data = self._load_excel_file(filename)

            self.file_loaded.emit(data)
            return True
        except Exception as e:
            QMessageBox.critical(parent_widget, "Error", f"Failed to open file:\n{str(e)}")
            return False

    def _load_csv_file(self, filename):
        """Загрузка данных из CSV файла"""
        with open(filename, 'r', encoding='utf-8') as file:
            return list(csv.reader(file))

    def _load_excel_file(self, filename):
        """Загрузка данных из Excel файла"""
        wb = load_workbook(filename)
        sheet = wb.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def save_file(self, parent_widget, data):
        """Сохранение данных в текущий файл"""
        if not self.current_file:
            return self.save_file_as(parent_widget, data)
            
        try:
            if self.current_file.endswith('.csv'):
                self._save_csv_data(self.current_file, data)
            else:
                self._save_excel_data(self.current_file, data)
            
            self.file_saved.emit(True)
            return True
        except Exception as e:
            QMessageBox.critical(parent_widget, "Error", f"Failed to save file:\n{str(e)}")
            return False

    def save_file_as(self, parent_widget, data):
        """Сохранение данных в новый файл"""
        filename, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Save File As",
            str(self.current_file) if self.current_file else "",
            "CSV Files (*.csv);;Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        
        if not filename:
            return False
            
        self.current_file = filename
        return self.save_file(parent_widget, data)

    def _save_csv_data(self, filename, data):
        """Сохранение данных в CSV файл"""
        with open(filename, 'w', encoding='utf-8', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)

    def _save_excel_data(self, filename, data):
        """Сохранение данных в Excel файл"""
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(filename)

    def import_excel(self, filename):
        """Импорт данных из Excel файла"""
        try:
            data = self._load_excel_file(filename)
            self.file_loaded.emit(data)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Excel Import Error", str(e))
            return False

    def import_sqlite(self, db_file, table_name):
        """Импорт данных из SQLite базы данных"""
        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            
            # Получаем данные
            cursor.execute(f"SELECT * FROM {table_name}")
            data = [list(row) for row in cursor.fetchall()]
            
            # Получаем названия колонок
            cursor.execute(f"PRAGMA table_info({table_name})")
            headers = [col[1] for col in cursor.fetchall()]
            
            # Добавляем заголовки как первую строку
            if headers:
                data.insert(0, headers)
            
            self.file_loaded.emit(data)
            return True
        except Exception as e:
            QMessageBox.critical(None, "SQLite Import Error", str(e))
            return False
        finally:
            if 'conn' in locals():
                conn.close()

    def export_to_excel(self, filename):
        """Экспорт данных в Excel файл"""
        if not self.model:
            return False
            
        try:
            wb = Workbook()
            ws = wb.active
            
            # Заголовки
            headers = [self.model.headerData(col, Qt.Horizontal) 
                      for col in range(self.model.columnCount())]
            ws.append(headers)
            
            # Данные
            for row in range(self.model.rowCount()):
                row_data = [self.model.data(self.model.index(row, col)) 
                           for col in range(self.model.columnCount())]
                ws.append(row_data)
            
            wb.save(filename)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Excel Export Error", str(e))
            return False

    def export_to_markdown(self, filename):
        """Экспорт данных в Markdown формат"""
        if not self.model:
            return False
            
        try:
            # Заголовки таблицы
            headers = [self.model.headerData(col, Qt.Horizontal) 
                      for col in range(self.model.columnCount())]
            md_content = "| " + " | ".join(headers) + " |\n"
            
            # Разделитель
            md_content += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            
            # Данные
            for row in range(self.model.rowCount()):
                row_data = [str(self.model.data(self.model.index(row, col)) or "")
                           for col in range(self.model.columnCount())]
                md_content += "| " + " | ".join(row_data) + " |\n"
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(md_content)
            
            return True
        except Exception as e:
            QMessageBox.critical(None, "Markdown Export Error", str(e))
            return False
    def export_to_csv(self, filename):
        """Экспорт данных в CSV файл"""
        if not self.model:
            return False
            
        try:
            with open(filename, 'w', encoding='utf-8', newline='') as file:
                writer = csv.writer(file)
                
                # Заголовки
                headers = [self.model.headerData(col, Qt.Horizontal) 
                          for col in range(self.model.columnCount())]
                writer.writerow(headers)
                
                # Данные
                for row in range(self.model.rowCount()):
                    row_data = [self.model.data(self.model.index(row, col)) 
                              for col in range(self.model.columnCount())]
                    writer.writerow(row_data)
            
            return True
        except Exception as e:
            QMessageBox.critical(None, "CSV Export Error", str(e))
            return False
